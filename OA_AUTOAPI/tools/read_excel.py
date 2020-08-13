import openpyxl
import os
from conf import project_path
from tools.read_config import ReadConfig

class ReadExel():
    case_path = os.path.join(project_path.test_config_path,"case.config")
    def __init__(self,file_name):
        self.file_name = file_name
        # self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        # self.sheet = self.workbook[sheet_name]
        # self.max_row = self.sheet.max_row
        # self.max_column = self.sheet.max_column
        self.mode = eval(ReadConfig().read_config(self.case_path,"MODE","mode")) #测试用例的匹配模式，运行哪些用例的哪些条

    def get_data(self):
        test_data = []
        #key是表名
        for key in self.mode:
            # ex = ReadExel(self.file_name)
            sheet = self.workbook[key]
            # print(key)
            # print(sheet.max_row)
            if self.mode[key] == "all":
                for i in range(2,sheet.max_row+1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i,1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['method'] = sheet.cell(i, 3).value
                    row_data['url'] = sheet.cell(i, 4).value
                    row_data['data'] = sheet.cell(i, 5).value
                    row_data['expected'] = sheet.cell(i, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
            else:
                for case_id in self.mode[key]:
                    row_data = {}
                    i = case_id+1
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['method'] = sheet.cell(i, 3).value
                    row_data['url'] = sheet.cell(i, 4).value
                    row_data['data'] = sheet.cell(i, 5).value
                    row_data['expected'] = sheet.cell(i, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)

        return test_data

    def write_data(self,id,sheet_name,result,test_result=None):
        sheet = self.workbook[sheet_name]
        sheet.cell(id,7).value = result
        sheet.cell(id,8).value = test_result
        self.workbook.save(self.file_name)

if __name__ == '__main__':
    excel_path = os.path.join(project_path.test_data_path,"test.xlsx")
    ex = ReadExel(excel_path)
    data =ex .get_data()
    print(type(data))
    print(len(data))
    print(data)

