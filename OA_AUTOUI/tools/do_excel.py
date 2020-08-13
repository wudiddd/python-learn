import openpyxl


class DoExcel():
    #excle三个对象：workbook，sheet，cell
    #openpyxl试从1开始的，不是从0开始的
    def __init__(self,fileName,sheetName):
        self.fileName = fileName
        # 打开文件
        self.excel = openpyxl.load_workbook(fileName)
        # 获取sheet
        self.table = self.excel[sheetName]
        # self.table = self.excel.get_sheet_by_name(sheetName) #好像已经过时了
        #行和列两个可迭代的生成器
        self.rows = self.table.rows
        self.column = self.table.columns
        #获取行数和列数
        self.max_row = self.table.max_row
        self.max_col = self.table.max_column
    def read_excel(self):
        # current_sheet = self.excel.active
        # print(current_sheet)
        row_datas = []
        #按行得到所有单元格
        rows_data = list(self.rows)
        #保存第一行表头到title列表
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        #读取用例数据并和标题聚合成字典
        for row_data in rows_data[1:]:
            datas = []
            for data in row_data:
                datas.append(data.value)
            row_datas.append(dict(zip(titles,datas)))
        return row_datas

    #写结果
    def write_excel(self,i,result):
        self.table.cell(i+2,6).value = result
        self.excel.save(self.fileName)



if __name__ == '__main__':
    data_file = "../TestData/login_data.xlsx"
    d = DoExcel(data_file, "login")
    print(d.read_excel())
    for i in range(d.max_row-1):
        d.write_excel(i,300)

